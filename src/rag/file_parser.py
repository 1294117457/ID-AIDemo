"""文件解析 - 对应 TS: rag/"""
import os
from typing import Optional


async def parse_file_to_text(file_path: str, extension: str) -> str:
    """
    解析文件为文本
    
    对应 TS: parseFileToText
    """
    ext = extension.lower().lstrip('.')
    
    parsers = {
        "pdf": _parse_pdf,
        "docx": _parse_docx,
        "doc": _parse_doc,
        "xlsx": _parse_xlsx,
        "xls": _parse_xlsx,
        "txt": _parse_txt,
    }
    
    parser = parsers.get(ext)
    if not parser:
        return f"[不支持的文件格式: {extension}]"
    
    try:
        return await parser(file_path)
    except Exception as e:
        print(f"[file_parser] parse error: {e}")
        return f"[文件解析失败: {str(e)}]"


async def _parse_pdf(file_path: str) -> str:
    """解析 PDF"""
    import pdfplumber
    
    text_parts = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)
    
    return "\n".join(text_parts)


async def _parse_docx(file_path: str) -> str:
    """解析 DOCX"""
    from docx import Document
    
    doc = Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs]
    return "\n".join(paragraphs)


async def _parse_doc(file_path: str) -> str:
    """解析 DOC (旧格式)"""
    import subprocess
    
    # 使用 textutil (macOS) 或 antiword (Linux)
    try:
        result = subprocess.run(
            ["textutil", "-convert", "txt", "-stdout", file_path],
            capture_output=True,
            text=True,
            timeout=30
        )
        if result.returncode == 0:
            return result.stdout
    except Exception:
        pass
    
    # Fallback: 尝试 mammoth (需要 docx 转换)
    return "[DOC 格式暂不支持，请另存为 DOCX 或 PDF]"


async def _parse_xlsx(file_path: str) -> str:
    """解析 Excel"""
    import openpyxl
    
    text_parts = []
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        text_parts.append(f"=== Sheet: {sheet_name} ===")
        
        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip():
                text_parts.append(row_text)
    
    return "\n".join(text_parts)


async def _parse_txt(file_path: str) -> str:
    """解析纯文本"""
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        return f.read()


def decode_filename(filename: str) -> str:
    """
    解码文件名 (处理中文)
    
    对应 TS: decodeFileName
    """
    # 尝试多种编码
    encodings = ['utf-8', 'gbk', 'gb2312', 'latin1']
    
    for encoding in encodings:
        try:
            return filename.encode(encoding).decode(encoding)
        except (UnicodeDecodeError, UnicodeEncodeError):
            continue
    
    # 最终 fallback
    return filename
